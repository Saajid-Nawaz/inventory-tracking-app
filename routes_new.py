"""
Enhanced Routes for Multi-Site Inventory Management System
"""

import os
from flask import render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from datetime import datetime, date
import logging
from io import BytesIO

from app import app, db
from models_new import (
    User, Site, Material, StockLevel, Transaction, IssueRequest, BatchIssueRequest, 
    BatchIssueItem, StockAdjustment, FIFOBatch, StockTransferRequest
)
from inventory_service import InventoryService, ReportService
from report_generator import PDFReportGenerator, ExcelReportGenerator

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# Initialize default data
def initialize_default_data():
    """Initialize default sites, materials, and users"""
    try:
        # Create default sites
        if not Site.query.first():
            sites = [
                Site(name='Main Construction Site', location='Downtown District'),
                Site(name='North Warehouse', location='North Industrial Zone'),
                Site(name='South Depot', location='South Storage Complex')
            ]
            for site in sites:
                db.session.add(site)
        
        # Create default materials
        if not Material.query.first():
            materials = [
                Material(name='Portland Cement', unit='bags', description='50kg bags', cost_per_unit=12.50, minimum_level=100),
                Material(name='Reinforcement Steel', unit='kg', description='Grade 60 rebar', cost_per_unit=0.85, minimum_level=1000),
                Material(name='Concrete Blocks', unit='pieces', description='8x8x16 blocks', cost_per_unit=2.25, minimum_level=500),
                Material(name='Sand', unit='cubic meters', description='Fine construction sand', cost_per_unit=25.00, minimum_level=50),
                Material(name='Gravel', unit='cubic meters', description='Coarse aggregate', cost_per_unit=30.00, minimum_level=30),
                Material(name='Lumber 2x4', unit='pieces', description='8ft pressure treated', cost_per_unit=8.50, minimum_level=200),
                Material(name='Roofing Sheets', unit='pieces', description='Galvanized iron sheets', cost_per_unit=15.75, minimum_level=100),
                Material(name='PVC Pipes', unit='meters', description='4-inch diameter', cost_per_unit=5.50, minimum_level=500)
            ]
            for material in materials:
                db.session.add(material)
        
        # Create default users
        if not User.query.first():
            # Site Engineers
            engineer1 = User(username='engineer1', role='site_engineer')
            engineer1.set_password('engineer123')
            
            engineer2 = User(username='engineer2', role='site_engineer')
            engineer2.set_password('engineer123')
            
            # Storesmen
            storesman1 = User(username='storesman1', role='storesman', assigned_site_id=1)
            storesman1.set_password('store123')
            
            storesman2 = User(username='storesman2', role='storesman', assigned_site_id=2)
            storesman2.set_password('store123')
            
            storesman3 = User(username='storesman3', role='storesman', assigned_site_id=3)
            storesman3.set_password('store123')
            
            users = [engineer1, engineer2, storesman1, storesman2, storesman3]
            for user in users:
                db.session.add(user)
        
        # Add sample inventory data if none exists
        if not Transaction.query.first():
            from inventory_service import InventoryService
            
            # Add sample inventory to each site
            sample_materials = [
                {'material_id': 1, 'quantity': 150, 'unit_cost': 12.50},  # Portland Cement
                {'material_id': 2, 'quantity': 2000, 'unit_cost': 0.85},  # Reinforcement Steel
                {'material_id': 3, 'quantity': 800, 'unit_cost': 2.25},   # Concrete Blocks
                {'material_id': 4, 'quantity': 75, 'unit_cost': 25.00},   # Sand
                {'material_id': 5, 'quantity': 50, 'unit_cost': 30.00},   # Gravel
            ]
            
            for site_id in [1, 2, 3]:  # For each site
                for material in sample_materials:
                    InventoryService.receive_material(
                        site_id=site_id,
                        material_id=material['material_id'],
                        quantity=material['quantity'],
                        unit_cost=material['unit_cost'],
                        project_code='INITIAL_STOCK',
                        created_by=1,  # engineer1
                        notes='Initial stock setup'
                    )
        
        db.session.commit()
        logging.info("Default data initialized successfully")
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error initializing default data: {str(e)}")


# Initialize data when app starts
with app.app_context():
    initialize_default_data()


# Authentication Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'site_engineer':
            return redirect(url_for('site_engineer_dashboard'))
        elif current_user.role == 'storesman':
            return redirect(url_for('storesman_dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    # Get demo accounts for display
    demo_accounts = [
        {'username': 'engineer1', 'password': 'engineer123', 'role': 'Site Engineer'},
        {'username': 'engineer2', 'password': 'engineer123', 'role': 'Site Engineer'},
        {'username': 'storesman1', 'password': 'store123', 'role': 'Storesman (Main Construction Site)'},
        {'username': 'storesman2', 'password': 'store123', 'role': 'Storesman (North Warehouse)'},
        {'username': 'storesman3', 'password': 'store123', 'role': 'Storesman (South Depot)'}
    ]
    
    return render_template('login.html', demo_accounts=demo_accounts)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully', 'info')
    return redirect(url_for('login'))


# Site Engineer Dashboard and Functions
@app.route('/site_engineer')
@login_required
def site_engineer_dashboard():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    # Get all sites summary
    sites = Site.query.all()
    
    # Get pending requests
    pending_individual_requests = IssueRequest.query.filter_by(status='pending').count()
    pending_batch_requests = BatchIssueRequest.query.filter_by(status='pending').count()
    
    # Get low stock alerts
    low_stock_items = InventoryService.get_low_stock_items()
    
    # Get recent transactions
    recent_transactions = InventoryService.get_transaction_history()[:10]
    
    return render_template('site_engineer_dashboard.html',
                         sites=sites,
                         pending_individual_requests=pending_individual_requests,
                         pending_batch_requests=pending_batch_requests,
                         low_stock_items=low_stock_items,
                         recent_transactions=recent_transactions)


@app.route('/manage_users')
@login_required
def manage_users():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    sites = Site.query.all()
    
    return render_template('manage_users.html', users=users, sites=sites)


@app.route('/add_user', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        assigned_site_id = request.form.get('assigned_site_id') if request.form.get('assigned_site_id') else None
        
        # Check if username already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('manage_users'))
        
        user = User(
            username=username,
            role=role,
            assigned_site_id=assigned_site_id
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'User {username} created successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating user: {str(e)}")
        flash('Error creating user', 'error')
    
    return redirect(url_for('manage_users'))


@app.route('/add_site', methods=['POST'])
@login_required
def add_site():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site = Site(
            name=request.form['name'],
            code=request.form.get('code', f"SITE{Site.query.count() + 1:03d}"),
            location=request.form.get('location'),
            is_active=True
        )
        db.session.add(site)
        db.session.commit()
        flash('Site added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error adding site: {str(e)}")
        flash('Error adding site', 'error')
    
    # Redirect to appropriate page based on referrer
    if 'system_settings' in request.referrer:
        return redirect(url_for('system_settings'))
    else:
        return redirect(url_for('manage_sites'))


@app.route('/edit_site', methods=['POST'])
@login_required
def edit_site():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site = Site.query.get(request.form['site_id'])
        site.name = request.form['name']
        site.location = request.form.get('location')
        site.updated_at = datetime.utcnow()
        
        # Set code if provided, otherwise keep existing
        if request.form.get('code'):
            site.code = request.form['code']
        
        if request.form.get('is_active'):
            site.is_active = bool(request.form.get('is_active'))
            
        db.session.commit()
        flash('Site updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating site: {str(e)}")
        flash('Error updating site', 'error')
    
    # Redirect to appropriate page based on referrer
    if 'system_settings' in request.referrer:
        return redirect(url_for('system_settings'))
    else:
        return redirect(url_for('manage_sites'))


@app.route('/edit_material', methods=['POST'])
@login_required
def edit_material():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        material = Material.query.get(request.form['material_id'])
        material.name = request.form['name']
        material.sku = request.form['sku']
        material.category = request.form.get('category')
        material.unit = request.form['unit']
        material.minimum_level = float(request.form['minimum_level']) if request.form.get('minimum_level') else None
        material.description = request.form.get('description')
        material.is_active = bool(request.form.get('is_active'))
        db.session.commit()
        flash('Material updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating material: {str(e)}")
        flash('Error updating material', 'error')
    
    return redirect(url_for('material_management'))


@app.route('/manage_sites')
@login_required
def manage_sites():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    sites = Site.query.all()
    return render_template('manage_sites.html', sites=sites)


@app.route('/material_management')
@login_required
def material_management():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    materials = Material.query.all()
    return render_template('material_management.html', materials=materials)


@app.route('/system_settings', methods=['GET', 'POST'])
@login_required
def system_settings():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Handle settings update
        flash('Settings updated successfully', 'success')
        return redirect(url_for('system_settings'))
    
    users = User.query.all()
    sites = Site.query.all()
    
    # Get system statistics
    total_transactions = Transaction.query.count()
    active_users = User.query.count()
    today_transactions = Transaction.query.filter(
        Transaction.created_at >= datetime.now().date()
    ).count()
    pending_approvals = IssueRequest.query.filter_by(status='pending').count()
    low_stock_items = len(InventoryService.get_low_stock_items())
    
    # Calculate total inventory value
    total_inventory_value = db.session.query(db.func.sum(StockLevel.total_value)).scalar() or 0
    
    return render_template('system_settings.html',
                         users=users,
                         sites=sites,
                         total_transactions=total_transactions,
                         active_users=active_users,
                         today_transactions=today_transactions,
                         pending_approvals=pending_approvals,
                         low_stock_items=low_stock_items,
                         total_inventory_value=total_inventory_value,
                         database_size=25,  # Placeholder
                         last_backup=None,  # Placeholder
                         settings={'company_name': 'Construction Company', 'currency': 'USD', 'default_tax_rate': 0, 'low_stock_threshold': 20})


@app.route('/manage_materials')
@login_required
def manage_materials():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    materials = Material.query.all()
    
    # Add missing fields for template compatibility
    for material in materials:
        material.sku = f"SKU-{material.id:04d}"
        material.category = "Construction Material"
        material.is_active = True
    
    return render_template('material_management.html', materials=materials)


@app.route('/add_material', methods=['POST'])
@login_required
def add_material():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        name = request.form.get('name')
        unit = request.form.get('unit')
        description = request.form.get('description')
        cost_per_unit = float(request.form.get('cost_per_unit', 0))
        minimum_level = float(request.form.get('minimum_level', 0))
        
        material = Material(
            name=name,
            unit=unit,
            description=description,
            cost_per_unit=cost_per_unit,
            minimum_level=minimum_level
        )
        
        db.session.add(material)
        db.session.commit()
        
        flash(f'Material {name} added successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error adding material: {str(e)}")
        flash('Error adding material', 'error')
    
    return redirect(url_for('manage_materials'))


@app.route('/download_material_template')
@login_required
def download_material_template():
    """Download Excel template for material upload"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create sample data for the template
        template_data = {
            'Material Name': [
                'Portland Cement',
                'Steel Rebar 12mm',
                'Ready Mix Concrete',
                'Electrical Wire 2.5mm',
                'PVC Pipe 110mm'
            ],
            'Unit': [
                'bags',
                'kg',
                'm3',
                'm',
                'm'
            ],
            'Description': [
                'High quality Portland cement for construction',
                'Steel reinforcement bars 12mm diameter',
                'Ready mix concrete M20 grade',
                'Electrical wire 2.5mm copper',
                'PVC drainage pipe 110mm diameter'
            ],
            'Cost per Unit': [
                8.50,
                0.85,
                120.00,
                2.30,
                15.75
            ],
            'Minimum Level': [
                50,
                500,
                10,
                100,
                20
            ],
            'Category': [
                'Cement',
                'Steel',
                'Concrete',
                'Electrical',
                'Plumbing'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Materials')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Materials']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='material_upload_template.xlsx'
        )
        
    except Exception as e:
        logging.error(f"Error creating template: {str(e)}")
        flash('Error creating template file', 'error')
        return redirect(url_for('manage_materials'))


@app.route('/upload_materials_excel', methods=['POST'])
@login_required
def upload_materials_excel():
    """Handle Excel file upload for bulk material import"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        import pandas as pd
        from werkzeug.utils import secure_filename
        
        # Check if file was uploaded
        if 'excel_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('manage_materials'))
        
        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('manage_materials'))
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Invalid file format. Please upload .xlsx or .xls files only.', 'error')
            return redirect(url_for('manage_materials'))
        
        # Read Excel file
        skip_header = request.form.get('skip_header') == '1'
        update_existing = request.form.get('update_existing') == '1'
        
        # Read the Excel file
        df = pd.read_excel(file, header=0 if skip_header else None)
        
        # Define expected columns
        expected_columns = ['Material Name', 'Unit', 'Description', 'Cost per Unit', 'Minimum Level', 'Category']
        
        # If no header, assign column names
        if not skip_header:
            df.columns = expected_columns[:len(df.columns)]
        
        # Validate required columns
        if 'Material Name' not in df.columns or 'Unit' not in df.columns:
            flash('Excel file must contain "Material Name" and "Unit" columns', 'error')
            return redirect(url_for('manage_materials'))
        
        # Process each row
        added_count = 0
        updated_count = 0
        error_count = 0
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row['Material Name']) or row['Material Name'].strip() == '':
                    continue
                
                material_name = str(row['Material Name']).strip()
                unit = str(row['Unit']).strip()
                
                # Validate unit
                valid_units = ['kg', 'bags', 'tons', 'm3', 'm2', 'm', 'pcs', 'liters']
                if unit.lower() not in valid_units:
                    logging.warning(f"Invalid unit '{unit}' for material '{material_name}'. Skipping.")
                    error_count += 1
                    continue
                
                # Check if material exists
                existing_material = Material.query.filter_by(name=material_name).first()
                
                if existing_material and not update_existing:
                    logging.info(f"Material '{material_name}' already exists. Skipping.")
                    continue
                
                # Prepare material data
                material_data = {
                    'name': material_name,
                    'unit': unit,
                    'description': str(row.get('Description', '')).strip() if pd.notna(row.get('Description')) else '',
                    'cost_per_unit': float(row.get('Cost per Unit', 0)) if pd.notna(row.get('Cost per Unit')) else 0.0,
                    'minimum_level': float(row.get('Minimum Level', 0)) if pd.notna(row.get('Minimum Level')) else 0.0
                }
                
                if existing_material and update_existing:
                    # Update existing material
                    existing_material.unit = material_data['unit']
                    existing_material.description = material_data['description']
                    existing_material.cost_per_unit = material_data['cost_per_unit']
                    existing_material.minimum_level = material_data['minimum_level']
                    updated_count += 1
                else:
                    # Create new material
                    new_material = Material(**material_data)
                    db.session.add(new_material)
                    added_count += 1
                    
            except Exception as e:
                logging.error(f"Error processing row {index}: {str(e)}")
                error_count += 1
                continue
        
        # Commit all changes
        db.session.commit()
        
        # Show success message
        success_msg = f"Excel upload completed! Added: {added_count}, Updated: {updated_count}"
        if error_count > 0:
            success_msg += f", Errors: {error_count}"
        
        flash(success_msg, 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error uploading materials: {str(e)}")
        flash(f'Error uploading materials: {str(e)}', 'error')
    
    return redirect(url_for('manage_materials'))


@app.route('/view_stock')
@login_required
def view_stock():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = request.args.get('site_id', type=int)
    
    try:
        stock_summary = InventoryService.get_stock_summary(site_id)
    except Exception as e:
        logging.error(f"Error getting stock summary: {str(e)}")
        stock_summary = []
        flash('Error loading stock data. Please try again.', 'error')
    
    sites = Site.query.all()
    selected_site = Site.query.get(site_id) if site_id else None
    
    return render_template('view_stock.html',
                         stock_summary=stock_summary,
                         sites=sites,
                         selected_site=selected_site)


@app.route('/download_stock_excel')
@login_required
def download_stock_excel():
    """Download stock levels as Excel file"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = request.args.get('site_id', type=int)
    
    try:
        # Get stock summary data
        stock_summary = InventoryService.get_stock_summary(site_id)
        
        # Create Excel file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        
        # Set title
        site_name = Site.query.get(site_id).name if site_id else 'All Sites'
        ws.title = f"Stock Levels - {site_name}"
        
        # Headers
        headers = ['Site', 'Material Name', 'Current Stock', 'Unit', 'Minimum Level', 'Status', 'Total Value', 'Last Updated']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_num, item in enumerate(stock_summary, 2):
            ws.cell(row=row_num, column=1, value=item.site_name)
            ws.cell(row=row_num, column=2, value=item.material_name)
            ws.cell(row=row_num, column=3, value=float(item.quantity))
            ws.cell(row=row_num, column=4, value=item.unit)
            ws.cell(row=row_num, column=5, value=float(item.minimum_level) if item.minimum_level else 0)
            
            # Status
            if item.minimum_level and item.quantity <= item.minimum_level:
                status = "Low Stock"
                status_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif item.minimum_level and item.quantity <= (item.minimum_level * 1.5):
                status = "Warning"
                status_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            else:
                status = "Normal"
                status_fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
            
            status_cell = ws.cell(row=row_num, column=6, value=status)
            status_cell.fill = status_fill
            
            ws.cell(row=row_num, column=7, value=float(item.total_value))
            ws.cell(row=row_num, column=8, value=item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else 'N/A')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"stock_levels_{site_name.replace(' ', '_')}_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"Error generating stock Excel: {str(e)}")
        flash('Error generating Excel file', 'error')
        return redirect(url_for('view_stock'))


@app.route('/approve_requests')
@login_required
def approve_requests():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    # Get pending individual requests
    individual_requests = IssueRequest.query.filter_by(status='pending').all()
    
    # Get pending batch requests
    batch_requests = BatchIssueRequest.query.filter_by(status='pending').all()
    
    # Get pending stock transfer requests
    transfer_requests = StockTransferRequest.query.filter_by(status='pending').all()
    
    # Get recent decisions
    recent_individual = IssueRequest.query.filter(
        IssueRequest.status.in_(['approved', 'rejected'])
    ).order_by(IssueRequest.reviewed_at.desc()).limit(10).all()
    
    recent_batch = BatchIssueRequest.query.filter(
        BatchIssueRequest.status.in_(['approved', 'rejected'])
    ).order_by(BatchIssueRequest.reviewed_at.desc()).limit(10).all()
    
    recent_transfers = StockTransferRequest.query.filter(
        StockTransferRequest.status.in_(['approved', 'rejected'])
    ).order_by(StockTransferRequest.reviewed_at.desc()).limit(10).all()
    
    # Get stock levels for validation
    stock_levels = {}
    for site in Site.query.all():
        stock_levels[site.id] = {}
        for stock in StockLevel.query.filter_by(site_id=site.id).all():
            stock_levels[site.id][stock.material_id] = stock
    
    return render_template('approve_requests.html',
                         individual_requests=individual_requests,
                         batch_requests=batch_requests,
                         transfer_requests=transfer_requests,
                         recent_individual=recent_individual,
                         recent_batch=recent_batch,
                         recent_transfers=recent_transfers,
                         stock_levels=stock_levels)


@app.route('/process_individual_request', methods=['POST'])
@login_required
def process_individual_request():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        request_id = int(request.form.get('request_id'))
        action = request.form.get('action')
        review_notes = request.form.get('review_notes')
        
        InventoryService.process_issue_request(
            request_id=request_id,
            approved_by=current_user.id,
            action=action,
            review_notes=review_notes
        )
        
        flash(f'Request {action}d successfully', 'success')
        
    except Exception as e:
        logging.error(f"Error processing request: {str(e)}")
        flash('Error processing request', 'error')
    
    return redirect(url_for('approve_requests'))


@app.route('/process_batch_request', methods=['POST'])
@login_required
def process_batch_request():
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        batch_id = request.form.get('batch_id')
        action = request.form.get('action')
        review_notes = request.form.get('review_notes')
        
        InventoryService.process_batch_issue_request(
            batch_id=batch_id,
            approved_by=current_user.id,
            action=action,
            review_notes=review_notes
        )
        
        flash(f'Batch request {action}d successfully', 'success')
        
    except Exception as e:
        logging.error(f"Error processing batch request: {str(e)}")
        flash('Error processing batch request', 'error')
    
    return redirect(url_for('approve_requests'))


# Storesman Dashboard and Functions
@app.route('/storesman')
@login_required
def storesman_dashboard():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = current_user.assigned_site_id
    if not site_id:
        flash('No site assigned to user', 'error')
        return redirect(url_for('logout'))
    
    # Get site information
    site = Site.query.get(site_id)
    
    # Get stock summary for assigned site
    stock_summary = InventoryService.get_stock_summary(site_id)
    
    # Get low stock items for the site
    low_stock_items = InventoryService.get_low_stock_items(site_id)
    
    # Get recent transactions for the site
    recent_transactions = InventoryService.get_transaction_history(site_id)[:10]
    
    # Get pending requests
    pending_individual = IssueRequest.query.filter_by(
        site_id=site_id,
        requested_by=current_user.id,
        status='pending'
    ).count()
    
    pending_batch = BatchIssueRequest.query.filter_by(
        site_id=site_id,
        requested_by=current_user.id,
        status='pending'
    ).count()
    
    return render_template('storesman_dashboard.html',
                         site=site,
                         stock_summary=stock_summary,
                         low_stock_items=low_stock_items,
                         recent_transactions=recent_transactions,
                         pending_individual=pending_individual,
                         pending_batch=pending_batch,
                         now=datetime.now())


@app.route('/receive_materials')
@login_required
def receive_materials():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    materials = Material.query.all()
    return render_template('receive_materials.html', materials=materials)


@app.route('/bulk_receive_materials')
@login_required
def bulk_receive_materials():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    materials = Material.query.all()
    return render_template('bulk_receive_materials.html', materials=materials)


@app.route('/process_bulk_receive_material', methods=['POST'])
@login_required
def process_bulk_receive_material():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site_id = current_user.assigned_site_id
        supplier = request.form.get('supplier')
        invoice_number = request.form.get('invoice_number')
        project_code = request.form.get('project_code')
        notes = request.form.get('notes')
        
        # Handle file upload
        supporting_document_url = None
        if 'supporting_document' in request.files:
            file = request.files['supporting_document']
            if file and file.filename:
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"bulk_{timestamp}_{filename}"
                file_path = os.path.join('uploads', filename)
                
                # Ensure uploads directory exists
                os.makedirs('uploads', exist_ok=True)
                file.save(file_path)
                supporting_document_url = file_path
        
        # Process multiple materials
        material_ids = request.form.getlist('material_id[]')
        quantities = request.form.getlist('quantity[]')
        unit_costs = request.form.getlist('unit_cost[]')
        
        transactions = []
        for i, material_id in enumerate(material_ids):
            if material_id and quantities[i] and unit_costs[i]:
                transaction = InventoryService.receive_material(
                    site_id=site_id,
                    material_id=int(material_id),
                    quantity=float(quantities[i]),
                    unit_cost=float(unit_costs[i]),
                    project_code=project_code,
                    created_by=current_user.id,
                    notes=f"Bulk receipt from {supplier} - Invoice: {invoice_number}. {notes}"
                )
                
                # Update transaction with supporting document URL
                if supporting_document_url:
                    transaction.supporting_document_url = supporting_document_url
                    
                transactions.append(transaction)
        
        db.session.commit()
        
        flash(f'Bulk material receipt processed successfully. {len(transactions)} transactions created.', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error processing bulk receipt: {str(e)}")
        flash(f'Error processing bulk receipt: {str(e)}', 'error')
    
    return redirect(url_for('bulk_receive_materials'))


@app.route('/view_document/<path:filename>')
@login_required
def view_document(filename):
    """View uploaded supporting documents"""
    try:
        file_path = os.path.join('uploads', filename)
        if os.path.exists(file_path):
            return send_file(file_path)
        else:
            flash('Document not found', 'error')
            return redirect(url_for('storesman_dashboard'))
    except Exception as e:
        logging.error(f"Error viewing document: {str(e)}")
        flash('Error accessing document', 'error')
        return redirect(url_for('storesman_dashboard'))

@app.route('/download_document/<path:filename>')
@login_required
def download_document(filename):
    """Download uploaded supporting documents"""
    try:
        file_path = os.path.join('uploads', filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            flash('Document not found', 'error')
            return redirect(url_for('storesman_dashboard'))
    except Exception as e:
        logging.error(f"Error downloading document: {str(e)}")
        flash('Error downloading document', 'error')
        return redirect(url_for('storesman_dashboard'))

@app.route('/transaction_documents/<int:transaction_id>')
@login_required
def transaction_documents(transaction_id):
    """View documents for a specific transaction"""
    try:
        transaction = Transaction.query.get_or_404(transaction_id)
        
        # Check access permissions
        if current_user.role == 'storesman' and transaction.site_id != current_user.assigned_site_id:
            flash('Access denied', 'error')
            return redirect(url_for('storesman_dashboard'))
        
        documents = []
        if transaction.supporting_document_url:
            filename = os.path.basename(transaction.supporting_document_url)
            documents.append({
                'filename': filename,
                'path': transaction.supporting_document_url,
                'exists': os.path.exists(transaction.supporting_document_url)
            })
        
        return render_template('transaction_documents.html', 
                             transaction=transaction, 
                             documents=documents)
    except Exception as e:
        logging.error(f"Error viewing transaction documents: {str(e)}")
        flash('Error accessing transaction documents', 'error')
        return redirect(url_for('storesman_dashboard'))


@app.route('/process_receive_material', methods=['POST'])
@login_required
def process_receive_material():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site_id = current_user.assigned_site_id
        material_id = int(request.form.get('material_id'))
        quantity = float(request.form.get('quantity'))
        unit_cost = float(request.form.get('unit_cost'))
        project_code = request.form.get('project_code')
        notes = request.form.get('notes')
        
        # Handle file upload
        supporting_document_url = None
        if 'supporting_document' in request.files:
            file = request.files['supporting_document']
            if file and file.filename:
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{filename}"
                file_path = os.path.join('uploads', filename)
                
                # Ensure uploads directory exists
                os.makedirs('uploads', exist_ok=True)
                file.save(file_path)
                supporting_document_url = file_path
        
        transaction = InventoryService.receive_material(
            site_id=site_id,
            material_id=material_id,
            quantity=quantity,
            unit_cost=unit_cost,
            project_code=project_code,
            created_by=current_user.id,
            notes=notes
        )
        
        # Update transaction with supporting document URL
        if supporting_document_url:
            transaction.supporting_document_url = supporting_document_url
            db.session.commit()
        
        flash(f'Material received successfully. Transaction: {transaction.serial_number}', 'success')
        
    except Exception as e:
        logging.error(f"Error receiving material: {str(e)}")
        flash(f'Error receiving material: {str(e)}', 'error')
    
    return redirect(url_for('receive_materials'))


@app.route('/request_materials')
@login_required
def request_materials():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = current_user.assigned_site_id
    stock_summary = InventoryService.get_stock_summary(site_id)
    
    # Get user's recent requests
    recent_requests = IssueRequest.query.filter_by(
        site_id=site_id,
        requested_by=current_user.id
    ).order_by(IssueRequest.requested_at.desc()).limit(10).all()
    
    return render_template('request_materials.html',
                         stock_summary=stock_summary,
                         recent_requests=recent_requests)


@app.route('/submit_material_request', methods=['POST'])
@login_required
def submit_material_request():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site_id = current_user.assigned_site_id
        material_id = int(request.form.get('material_id'))
        quantity_requested = float(request.form.get('quantity_requested'))
        project_code = request.form.get('project_code')
        purpose = request.form.get('purpose')
        
        # Check if sufficient stock is available
        stock_level = StockLevel.query.filter_by(site_id=site_id, material_id=material_id).first()
        if not stock_level or stock_level.quantity < quantity_requested:
            flash('Insufficient stock available for this request', 'error')
            return redirect(url_for('request_materials'))
        
        issue_request = IssueRequest(
            site_id=site_id,
            material_id=material_id,
            quantity_requested=quantity_requested,
            project_code=project_code,
            purpose=purpose,
            requested_by=current_user.id
        )
        
        db.session.add(issue_request)
        db.session.commit()
        
        flash('Material request submitted successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error submitting request: {str(e)}")
        flash('Error submitting request', 'error')
    
    return redirect(url_for('request_materials'))


@app.route('/batch_request')
@login_required
def batch_request():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = current_user.assigned_site_id
    stock_summary = InventoryService.get_stock_summary(site_id)
    
    # Get user's recent batch requests
    recent_batch_requests = BatchIssueRequest.query.filter_by(
        site_id=site_id,
        requested_by=current_user.id
    ).order_by(BatchIssueRequest.requested_at.desc()).limit(10).all()
    
    return render_template('batch_request.html',
                         stock_summary=stock_summary,
                         recent_batch_requests=recent_batch_requests)


@app.route('/submit_batch_request', methods=['POST'])
@login_required
def submit_batch_request():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site_id = current_user.assigned_site_id
        project_code = request.form.get('project_code')
        purpose = request.form.get('purpose')
        
        # Generate batch ID
        batch_id = BatchIssueRequest.generate_batch_id()
        
        # Create batch request
        batch_request = BatchIssueRequest(
            batch_id=batch_id,
            site_id=site_id,
            project_code=project_code,
            purpose=purpose,
            requested_by=current_user.id
        )
        
        db.session.add(batch_request)
        db.session.flush()  # Get the batch_id
        
        # Process batch items
        item_count = 0
        i = 0
        while request.form.get(f'material_id_{i}'):
            material_id = int(request.form.get(f'material_id_{i}'))
            quantity = float(request.form.get(f'quantity_{i}'))
            
            if quantity > 0:
                # Check stock availability
                stock_level = StockLevel.query.filter_by(site_id=site_id, material_id=material_id).first()
                if not stock_level or stock_level.quantity < quantity:
                    material = Material.query.get(material_id)
                    flash(f'Insufficient stock for {material.name}', 'error')
                    return redirect(url_for('batch_request'))
                
                batch_item = BatchIssueItem(
                    batch_id=batch_id,
                    material_id=material_id,
                    quantity_requested=quantity
                )
                db.session.add(batch_item)
                item_count += 1
            
            i += 1
        
        if item_count == 0:
            flash('Please add at least one material to the batch request', 'error')
            return redirect(url_for('batch_request'))
        
        db.session.commit()
        flash(f'Batch request {batch_id} submitted successfully', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error submitting batch request: {str(e)}")
        flash('Error submitting batch request', 'error')
    
    return redirect(url_for('batch_request'))


@app.route('/view_transactions')
@login_required
def view_transactions():
    """View all transactions for site engineers"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    # Get filter parameters
    site_id = request.args.get('site_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    transaction_type = request.args.get('type')
    
    # Convert date strings to datetime objects
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
    
    # Get transactions based on filters
    transactions = InventoryService.get_transaction_history(
        site_id=site_id,
        start_date=start_date_obj,
        end_date=end_date_obj
    )
    
    # Filter by transaction type if specified
    if transaction_type:
        transactions = [txn for txn in transactions if txn.type == transaction_type]
    
    # Get sites for filter dropdown
    sites = Site.query.all()
    
    return render_template('view_transactions.html',
                         transactions=transactions,
                         sites=sites,
                         selected_site_id=site_id,
                         selected_start_date=start_date,
                         selected_end_date=end_date,
                         selected_type=transaction_type)

# Report Generation Routes
@app.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    sites = Site.query.all()
    user_site_id = current_user.assigned_site_id if current_user.role == 'storesman' else None
    
    if request.method == 'POST':
        # Handle report generation
        flash('Report generated successfully', 'success')
        return redirect(url_for('reports'))
    
    return render_template('reports.html', sites=sites, user_site_id=user_site_id)


@app.route('/generate_daily_report')
@login_required
def generate_daily_report():
    site_id = request.args.get('site_id', type=int)
    report_date = request.args.get('report_date')
    report_format = request.args.get('format', 'pdf')
    
    if not site_id or not report_date:
        flash('Site and date are required', 'error')
        return redirect(url_for('reports'))
    
    try:
        report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
        site = Site.query.get(site_id)
        
        if not site:
            flash('Site not found', 'error')
            return redirect(url_for('reports'))
        
        # Check access permissions
        if current_user.role == 'storesman' and current_user.assigned_site_id != site_id:
            flash('Access denied to this site', 'error')
            return redirect(url_for('reports'))
        
        issues_data = ReportService.generate_daily_issues_report(site_id, report_date)
        # Get current currency setting
        currency = 'ZMW'  # Default currency setting
        
        if report_format == 'excel':
            excel_data = ExcelReportGenerator.generate_daily_issues_excel(site.name, report_date, issues_data)
            return send_file(
                BytesIO(excel_data),
                as_attachment=True,
                download_name=f'daily_issues_{site.name}_{report_date.strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            pdf_data = PDFReportGenerator.generate_daily_issues_report(site.name, report_date, issues_data, currency)
            return send_file(
                BytesIO(pdf_data),
                as_attachment=True,
                download_name=f'daily_issues_{site.name}_{report_date.strftime("%Y%m%d")}.pdf',
                mimetype='application/pdf'
            )
        
    except Exception as e:
        logging.error(f"Error generating daily report: {str(e)}")
        flash('Error generating report', 'error')
        return redirect(url_for('reports'))


@app.route('/generate_stock_report')
@login_required
def generate_stock_report():
    site_id = request.args.get('site_id', type=int)
    format_type = request.args.get('format', 'pdf')
    
    if not site_id:
        flash('Site is required', 'error')
        return redirect(url_for('reports'))
    
    try:
        site = Site.query.get(site_id)
        
        if not site:
            flash('Site not found', 'error')
            return redirect(url_for('reports'))
        
        # Check access permissions
        if current_user.role == 'storesman' and current_user.assigned_site_id != site_id:
            flash('Access denied to this site', 'error')
            return redirect(url_for('reports'))
        
        stock_data = ReportService.generate_stock_summary_report(site_id)
        # Debug logging
        logging.info(f"Stock data retrieved for site {site_id}: {len(stock_data)} items")
        if stock_data:
            for i, item in enumerate(stock_data[:3]):  # Log first 3 items
                logging.info(f"Stock item {i+1}: {item.material_name} - {item.quantity} {item.unit}")
        
        # Get current currency setting
        currency = 'ZMW'  # Updated to ZMW
        
        if format_type == 'excel':
            excel_data = ExcelReportGenerator.generate_stock_summary_excel(site.name, stock_data)
            return send_file(
                BytesIO(excel_data),
                as_attachment=True,
                download_name=f'stock_summary_{site.name}_{datetime.now().strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            pdf_data = PDFReportGenerator.generate_stock_summary_report(site.name, stock_data, currency)
            return send_file(
                BytesIO(pdf_data),
                as_attachment=True,
                download_name=f'stock_summary_{site.name}_{datetime.now().strftime("%Y%m%d")}.pdf',
                mimetype='application/pdf'
            )
        
    except Exception as e:
        logging.error(f"Error generating stock report: {str(e)}")
        flash('Error generating report', 'error')
        return redirect(url_for('reports'))


@app.route('/generate_transaction_history_report')
@login_required
def generate_transaction_history_report():
    """Generate transaction history report"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = request.args.get('site_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    report_format = request.args.get('format', 'pdf')  # Default to PDF
    
    logging.info(f"Transaction history report requested for site_id={site_id}, start_date={start_date}, end_date={end_date}, format={report_format}")
    
    if not site_id:
        flash('Site selection is required', 'error')
        return redirect(url_for('reports'))
    
    try:
        # Parse dates if provided
        start_date_obj = None
        end_date_obj = None
        
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        site = Site.query.get(site_id)
        if not site:
            flash('Site not found', 'error')
            return redirect(url_for('reports'))
        
        # Get transaction history
        transactions_data = ReportService.generate_transaction_history_report(
            site_id, start_date_obj, end_date_obj
        )
        
        # Debug logging
        logging.info(f"Transaction data for site {site_id}: {len(transactions_data)} transactions")
        if transactions_data:
            for i, txn in enumerate(transactions_data[:3]):  # Log first 3 transactions
                logging.info(f"Transaction {i+1}: {txn.serial_number} - {txn.type} - {txn.quantity}")
        
        # Generate filename
        date_suffix = f"_{start_date}_{end_date}" if start_date and end_date else ""
        
        if report_format == 'excel':
            # Generate Excel report
            excel_data = ExcelReportGenerator.generate_transaction_history_excel(
                site.name, transactions_data, start_date_obj, end_date_obj
            )
            
            filename = f"transaction_history_{site.name.replace(' ', '_')}{date_suffix}.xlsx"
            
            return send_file(
                BytesIO(excel_data),
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Generate PDF report
            pdf_data = PDFReportGenerator.generate_transaction_history_report(
                site.name, start_date_obj, end_date_obj, transactions_data
            )
            
            filename = f"transaction_history_{site.name.replace(' ', '_')}{date_suffix}.pdf"
            
            return send_file(
                BytesIO(pdf_data),
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )
        
    except Exception as e:
        logging.error(f"Error generating transaction history report: {str(e)}")
        flash('Error generating report', 'error')
        return redirect(url_for('reports'))


@app.route('/process_transfer_request', methods=['POST'])
@login_required
def process_transfer_request():
    """Process stock transfer request (approve/reject)"""
    if current_user.role != 'site_engineer':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        transfer_id = request.form.get('transfer_id')
        action = request.form.get('action')
        review_notes = request.form.get('review_notes')
        
        InventoryService.process_stock_transfer_request(
            transfer_id=transfer_id,
            approved_by=current_user.id,
            action=action,
            review_notes=review_notes
        )
        
        flash(f'Transfer request {action}d successfully', 'success')
        
    except Exception as e:
        logging.error(f"Error processing transfer request: {str(e)}")
        flash(f'Error processing transfer request: {str(e)}', 'error')
    
    return redirect(url_for('approve_requests'))


@app.route('/stock_transfer', methods=['GET', 'POST'])
@login_required
def stock_transfer():
    """Handle stock transfer requests"""
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            from_site_id = int(request.form.get('from_site_id'))
            to_site_id = int(request.form.get('to_site_id'))
            priority = request.form.get('priority', 'normal')
            reason = request.form.get('reason')
            
            # Validate sites
            if from_site_id == to_site_id:
                flash('Source and destination sites must be different', 'error')
                return redirect(url_for('stock_transfer'))
            
            # Check if storesman has access to from_site
            if current_user.assigned_site_id != from_site_id:
                flash('You can only transfer from your assigned site', 'error')
                return redirect(url_for('stock_transfer'))
            
            # Parse materials
            materials = []
            form_data = request.form.to_dict()
            
            # Extract material data from form
            for key, value in form_data.items():
                if key.startswith('materials[') and key.endswith('][material_id]'):
                    index = key.split('[')[1].split(']')[0]
                    material_id = int(value)
                    quantity_key = f'materials[{index}][quantity]'
                    
                    if quantity_key in form_data:
                        quantity = float(form_data[quantity_key])
                        materials.append({
                            'material_id': material_id,
                            'quantity': quantity
                        })
            
            if not materials:
                flash('At least one material must be selected', 'error')
                return redirect(url_for('stock_transfer'))
            
            # Create transfer request
            transfer_request = InventoryService.create_stock_transfer_request(
                from_site_id=from_site_id,
                to_site_id=to_site_id,
                materials=materials,
                requested_by=current_user.id,
                reason=reason,
                priority=priority
            )
            
            flash(f'Stock transfer request {transfer_request.transfer_id} submitted successfully', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            logging.error(f"Error creating stock transfer request: {str(e)}")
            flash('Error creating transfer request', 'error')
            return redirect(url_for('stock_transfer'))
    
    # GET request - show form
    sites = Site.query.all()
    materials = Material.query.all()
    
    return render_template('stock_transfer.html',
                         sites=sites,
                         materials=materials)


@app.route('/stock_adjustments')
@login_required
def stock_adjustments():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    site_id = current_user.assigned_site_id
    stock_summary = InventoryService.get_stock_summary(site_id)
    
    # Get recent adjustments
    recent_adjustments = StockAdjustment.query.filter_by(
        site_id=site_id,
        adjusted_by=current_user.id
    ).order_by(StockAdjustment.adjusted_at.desc()).limit(10).all()
    
    return render_template('stock_adjustments.html',
                         stock_summary=stock_summary,
                         recent_adjustments=recent_adjustments)


@app.route('/process_stock_adjustment', methods=['POST'])
@login_required
def process_stock_adjustment():
    if current_user.role != 'storesman':
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    try:
        site_id = current_user.assigned_site_id
        material_id = int(request.form.get('material_id'))
        physical_count = float(request.form.get('physical_count'))
        reason = request.form.get('reason')
        notes = request.form.get('notes', '')
        
        # Get current stock level
        current_stock = db.session.query(StockLevel).filter_by(
            site_id=site_id,
            material_id=material_id
        ).first()
        
        if not current_stock:
            flash('Material not found in stock', 'error')
            return redirect(url_for('stock_adjustments'))
        
        expected_quantity = current_stock.quantity
        
        # Use inventory service to adjust stock
        InventoryService.adjust_stock(
            site_id=site_id,
            material_id=material_id,
            expected_quantity=expected_quantity,
            actual_quantity=physical_count,
            reason=reason,
            adjusted_by=current_user.id
        )
        
        adjustment_type = 'increase' if physical_count > expected_quantity else 'decrease'
        adjustment_amount = abs(physical_count - expected_quantity)
        
        flash(f'Stock adjustment recorded: {adjustment_type} of {adjustment_amount:.2f} units', 'success')
        
    except Exception as e:
        logging.error(f"Error processing stock adjustment: {str(e)}")
        flash('Error processing stock adjustment', 'error')
    
    return redirect(url_for('stock_adjustments'))


# API Endpoints for AJAX requests
@app.route('/api/stock_levels/<int:site_id>')
@login_required
def api_stock_levels(site_id):
    # Check access permissions
    if current_user.role == 'storesman' and current_user.assigned_site_id != site_id:
        return jsonify({'error': 'Access denied'}), 403
    
    stock_summary = InventoryService.get_stock_summary(site_id)
    
    data = []
    for item in stock_summary:
        data.append({
            'material_id': item.material_id,
            'material_name': item.material_name,
            'unit': item.unit,
            'quantity': item.quantity,
            'total_value': item.total_value,
            'average_cost': item.total_value / item.quantity if item.quantity > 0 else 0,
            'minimum_level': item.minimum_level,
            'is_low_stock': item.quantity < item.minimum_level
        })
    
    return jsonify(data)


@app.route('/api/materials')
@login_required
def api_materials():
    materials = Material.query.all()
    
    data = []
    for material in materials:
        data.append({
            'id': material.id,
            'name': material.name,
            'unit': material.unit,
            'description': material.description,
            'cost_per_unit': material.cost_per_unit,
            'minimum_level': material.minimum_level
        })
    
    return jsonify(data)


@app.route('/api/sites')
@login_required
def api_sites():
    sites = Site.query.all()
    
    data = []
    for site in sites:
        data.append({
            'id': site.id,
            'name': site.name,
            'location': site.location,
            'created_at': site.created_at.isoformat()
        })
    
    return jsonify(data)


@app.route('/api/test_transactions/<int:site_id>')
@login_required
def test_transactions(site_id):
    """Test endpoint to debug transaction history"""
    try:
        from inventory_service import InventoryService
        transactions = InventoryService.get_transaction_history(site_id=site_id)
        
        data = []
        for txn in transactions:
            data.append({
                'id': txn.id,
                'serial_number': txn.serial_number,
                'type': txn.type,
                'quantity': txn.quantity,
                'site_name': txn.site_name if hasattr(txn, 'site_name') else 'Unknown',
                'material_name': txn.material_name if hasattr(txn, 'material_name') else 'Unknown',
                'created_at': txn.created_at.isoformat(),
                'creator': txn.creator_user.username if txn.creator_user else 'Unknown'
            })
        
        return jsonify({
            'count': len(data),
            'transactions': data[:5]  # Return first 5 for testing
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/pending_counts')
@login_required
def api_pending_counts():
    """API endpoint for dashboard pending counts"""
    if current_user.role == 'site_engineer':
        # Count pending individual requests
        pending_individual = IssueRequest.query.filter_by(status='pending').count()
        # Count pending batch requests
        pending_batch = BatchIssueRequest.query.filter_by(status='pending').count()
        total_pending = pending_individual + pending_batch
    else:
        # For storesmen, count their own pending requests
        pending_individual = IssueRequest.query.filter_by(
            created_by=current_user.id, status='pending'
        ).count()
        pending_batch = BatchIssueRequest.query.filter_by(
            created_by=current_user.id, status='pending'
        ).count()
        total_pending = pending_individual + pending_batch
    
    return jsonify({
        'pending_individual': pending_individual,
        'pending_batch': pending_batch,
        'total_pending': total_pending
    })


# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error_code=404, error_message="Page not found"), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('error.html', error_code=500, error_message="Internal server error"), 500


@app.errorhandler(403)
def forbidden(error):
    return render_template('error.html', error_code=403, error_message="Access forbidden"), 403