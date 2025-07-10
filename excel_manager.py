import os
import pandas as pd
from datetime import datetime
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelManager:
    def __init__(self):
        self.excel_file = os.path.join('data', 'construction_materials.xlsx')
        self.ensure_excel_file_exists()
    
    def ensure_excel_file_exists(self):
        """Create Excel file with proper structure if it doesn't exist"""
        if not os.path.exists(self.excel_file):
            try:
                wb = Workbook()
                
                # Create Current Stock sheet
                ws_stock = wb.active
                ws_stock.title = "Current Stock"
                stock_headers = ['Material Name', 'Quantity', 'Unit', 'Last Updated', 'Updated By']
                ws_stock.append(stock_headers)
                
                # Create Issuance Log sheet
                ws_log = wb.create_sheet("Issuance Log")
                log_headers = ['Date', 'Material Name', 'Quantity Issued', 'Unit', 'Issued By', 'Notes', 'Remaining Stock']
                ws_log.append(log_headers)
                
                wb.save(self.excel_file)
                logging.info(f"Created new Excel file: {self.excel_file}")
                
            except Exception as e:
                logging.error(f"Error creating Excel file: {str(e)}")
                raise e
    
    def get_current_stock(self):
        """Get current stock from Excel file"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Current Stock')
            
            # Convert to list of dictionaries
            stock_list = []
            for _, row in df.iterrows():
                stock_list.append({
                    'material_name': row['Material Name'],
                    'quantity': row['Quantity'],
                    'unit': row['Unit'],
                    'last_updated': row['Last Updated'],
                    'updated_by': row['Updated By']
                })
            
            return stock_list
            
        except FileNotFoundError:
            logging.warning("Excel file not found, creating new one")
            self.ensure_excel_file_exists()
            return []
        except Exception as e:
            logging.error(f"Error reading Excel file: {str(e)}")
            return []
    
    def update_stock(self, materials):
        """Update stock levels with new materials"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb['Current Stock']
            
            # Read current stock
            current_stock = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # If material name is not empty
                    material_name = row[0]
                    current_stock[material_name] = {
                        'quantity': row[1] or 0,
                        'unit': row[2] or 'units',
                        'last_updated': row[3],
                        'updated_by': row[4]
                    }
            
            # Update with new materials
            for material in materials:
                material_name = material['name']
                if material_name in current_stock:
                    # Add to existing stock
                    current_stock[material_name]['quantity'] += material['quantity']
                else:
                    # Add new material
                    current_stock[material_name] = {
                        'quantity': material['quantity'],
                        'unit': material['unit'],
                        'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'updated_by': 'Site Engineer'
                    }
                
                # Update timestamp
                current_stock[material_name]['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                current_stock[material_name]['updated_by'] = 'Site Engineer'
            
            # Clear existing data and write updated stock
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            
            # Write updated stock
            row_num = 2
            for material_name, data in current_stock.items():
                ws[f'A{row_num}'] = material_name
                ws[f'B{row_num}'] = data['quantity']
                ws[f'C{row_num}'] = data['unit']
                ws[f'D{row_num}'] = data['last_updated']
                ws[f'E{row_num}'] = data['updated_by']
                row_num += 1
            
            wb.save(self.excel_file)
            logging.info(f"Updated stock with {len(materials)} materials")
            
        except Exception as e:
            logging.error(f"Error updating stock: {str(e)}")
            raise e
    
    def record_issuance(self, material_name, quantity_issued, unit, issued_by, notes=''):
        """Record material issuance and update stock"""
        try:
            wb = load_workbook(self.excel_file)
            
            # Update current stock
            ws_stock = wb['Current Stock']
            updated = False
            
            for row in ws_stock.iter_rows(min_row=2, max_row=ws_stock.max_row):
                if row[0].value == material_name:
                    current_quantity = row[1].value or 0
                    new_quantity = current_quantity - quantity_issued
                    row[1].value = max(0, new_quantity)  # Don't allow negative stock
                    row[3].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    row[4].value = issued_by
                    updated = True
                    break
            
            if not updated:
                logging.warning(f"Material {material_name} not found in stock")
                remaining_stock = 0
            else:
                remaining_stock = max(0, current_quantity - quantity_issued)
            
            # Add to issuance log
            ws_log = wb['Issuance Log']
            new_row = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                material_name,
                quantity_issued,
                unit,
                issued_by,
                notes,
                remaining_stock
            ]
            ws_log.append(new_row)
            
            wb.save(self.excel_file)
            logging.info(f"Recorded issuance: {quantity_issued} {unit} of {material_name}")
            
        except Exception as e:
            logging.error(f"Error recording issuance: {str(e)}")
            raise e
    
    def get_issuance_log(self, limit=None):
        """Get issuance log from Excel file"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Issuance Log')
            
            if limit:
                df = df.tail(limit)
            
            # Convert to list of dictionaries
            log_list = []
            for _, row in df.iterrows():
                log_list.append({
                    'date': row['Date'],
                    'material_name': row['Material Name'],
                    'quantity_issued': row['Quantity Issued'],
                    'unit': row['Unit'],
                    'issued_by': row['Issued By'],
                    'notes': row['Notes'],
                    'remaining_stock': row['Remaining Stock']
                })
            
            return log_list
            
        except Exception as e:
            logging.error(f"Error reading issuance log: {str(e)}")
            return []
