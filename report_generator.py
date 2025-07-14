"""
Report Generator Service
Handles PDF and Excel export functionality
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import pandas as pd
from datetime import datetime
import os
import logging


class PDFReportGenerator:
    """Generate PDF reports for various inventory data"""
    
    @staticmethod
    def format_currency(amount, currency='USD'):
        """Format currency based on currency type"""
        if currency == 'ZMW':
            return f"K{amount:.2f}"
        elif currency == 'EUR':
            return f"€{amount:.2f}"
        elif currency == 'GBP':
            return f"£{amount:.2f}"
        elif currency == 'CAD':
            return f"C${amount:.2f}"
        else:  # Default to USD
            return f"${amount:.2f}"
    
    @staticmethod
    def generate_daily_issues_report(site_name, report_date, issues_data, currency='USD'):
        """
        Generate PDF report for daily material issues
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph(f"Daily Material Issues Report", title_style)
        story.append(title)
        
        # Site and date info
        site_info = Paragraph(f"<b>Site:</b> {site_name}<br/><b>Date:</b> {report_date.strftime('%B %d, %Y')}", styles['Normal'])
        story.append(site_info)
        story.append(Spacer(1, 20))
        
        if issues_data:
            # Create table data
            table_data = [
                ['Serial Number', 'Time', 'Material', 'Quantity', 'Unit Cost', 'Total Value', 'Project Code']
            ]
            
            total_value = 0
            for issue in issues_data:
                table_data.append([
                    issue.serial_number,
                    issue.created_at.strftime('%H:%M'),
                    issue.material_name,
                    f"{abs(issue.quantity)} {issue.unit}",
                    PDFReportGenerator.format_currency(issue.unit_cost, currency),
                    PDFReportGenerator.format_currency(abs(issue.total_value), currency),
                    issue.issued_to_project_code or 'N/A'
                ])
                total_value += abs(issue.total_value)
            
            # Add total row
            table_data.append(['', '', '', '', '', PDFReportGenerator.format_currency(total_value, currency), 'TOTAL'])
            
            # Create table
            table = Table(table_data, colWidths=[1.2*inch, 0.8*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No material issues recorded for this date.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        return pdf_data
    
    @staticmethod
    def generate_stock_summary_report(site_name, stock_data, currency='USD'):
        """
        Generate PDF report for stock summary
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph(f"Stock Summary Report", title_style)
        story.append(title)
        
        # Site and date info
        site_info = Paragraph(f"<b>Site:</b> {site_name}<br/><b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal'])
        story.append(site_info)
        story.append(Spacer(1, 20))
        
        if stock_data:
            # Create table data
            table_data = [
                ['Material Name', 'Unit', 'Quantity', 'Total Value', 'Avg Cost', 'Min Level', 'Status']
            ]
            
            total_inventory_value = 0
            low_stock_count = 0
            
            for item in stock_data:
                avg_cost = item.total_value / item.quantity if item.quantity > 0 else 0
                status = "LOW STOCK" if item.quantity < item.minimum_level else "OK"
                if item.quantity < item.minimum_level:
                    low_stock_count += 1
                
                table_data.append([
                    item.material_name,
                    item.unit,
                    f"{item.quantity:.2f}",
                    PDFReportGenerator.format_currency(item.total_value, currency),
                    PDFReportGenerator.format_currency(avg_cost, currency),
                    f"{item.minimum_level:.2f}",
                    status
                ])
                total_inventory_value += item.total_value
            
            # Create table
            table = Table(table_data, colWidths=[1.5*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            # Color code low stock items
            for i, item in enumerate(stock_data, 1):
                if item.quantity < item.minimum_level:
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.lightcoral),
                    ]))
            
            story.append(table)
            
            # Summary
            story.append(Spacer(1, 20))
            summary_text = f"""
            <b>Summary:</b><br/>
            Total Inventory Value: {PDFReportGenerator.format_currency(total_inventory_value, currency)}<br/>
            Total Materials: {len(stock_data)}<br/>
            Low Stock Items: {low_stock_count}
            """
            story.append(Paragraph(summary_text, styles['Normal']))
        else:
            story.append(Paragraph("No stock data available.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        return pdf_data

    @staticmethod
    def generate_transaction_history_report(site_name, start_date, end_date, transactions_data, currency='ZMW'):
        """
        Generate PDF report for transaction history
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"Transaction History Report - {site_name}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Date range
        date_range = ""
        if start_date and end_date:
            date_range = f"From {start_date} to {end_date}"
        elif start_date:
            date_range = f"From {start_date}"
        elif end_date:
            date_range = f"Until {end_date}"
        else:
            date_range = "All transactions"
        
        story.append(Paragraph(date_range, styles['Normal']))
        story.append(Spacer(1, 12))
        
        if transactions_data:
            # Create table headers
            table_data = [
                ['Date', 'Transaction ID', 'Type', 'Material', 'Quantity', 'Unit Cost', 'Total Value', 'Project Code', 'Created By']
            ]
            
            total_value = 0
            for transaction in transactions_data:
                created_by = transaction.creator_user.username if transaction.creator_user else 'Unknown'
                table_data.append([
                    transaction.created_at.strftime('%Y-%m-%d %H:%M'),
                    transaction.serial_number,
                    transaction.type.title(),
                    transaction.material_name,
                    f"{transaction.quantity:.2f}",
                    PDFReportGenerator.format_currency(transaction.unit_cost, currency),
                    PDFReportGenerator.format_currency(transaction.total_value, currency),
                    transaction.issued_to_project_code or 'N/A',
                    created_by
                ])
                total_value += transaction.total_value
            
            # Create table
            table = Table(table_data, colWidths=[1*inch, 1.2*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7)
            ]))
            
            # Color code transaction types
            for i, transaction in enumerate(transactions_data, 1):
                if transaction.type == 'issue':
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.lightcoral),
                    ]))
                elif transaction.type == 'receive':
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.lightgreen),
                    ]))
            
            story.append(table)
            
            # Summary
            story.append(Spacer(1, 20))
            summary_text = f"""
            <b>Summary:</b><br/>
            Total Transactions: {len(transactions_data)}<br/>
            Total Value Impact: {PDFReportGenerator.format_currency(total_value, currency)}<br/>
            Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            """
            story.append(Paragraph(summary_text, styles['Normal']))
        else:
            story.append(Paragraph("No transactions found for the selected period.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        return pdf_data


class ExcelReportGenerator:
    """Generate Excel reports for various inventory data"""
    
    @staticmethod
    def generate_stock_summary_excel(site_name, stock_data):
        """
        Generate Excel report for stock summary
        """
        buffer = BytesIO()
        
        # Create DataFrame
        data = []
        logging.info(f"Processing {len(stock_data)} stock items for Excel report")
        
        for item in stock_data:
            avg_cost = item.total_value / item.quantity if item.quantity > 0 else 0
            status = "LOW STOCK" if item.quantity < item.minimum_level else "OK"
            
            data.append({
                'Material Name': item.material_name,
                'Unit': item.unit,
                'Quantity': item.quantity,
                'Total Value': item.total_value,
                'Average Cost': avg_cost,
                'Minimum Level': item.minimum_level,
                'Status': status
            })
        
        df = pd.DataFrame(data)
        logging.info(f"Created DataFrame with {len(df)} rows")
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock Summary', index=False)
            
            # Create summary sheet
            summary_data = {
                'Metric': ['Total Inventory Value', 'Total Materials', 'Low Stock Items'],
                'Value': [
                    df['Total Value'].sum(),
                    len(df),
                    len(df[df['Status'] == 'LOW STOCK'])
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format stock summary sheet
            worksheet = writer.sheets['Stock Summary']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data
    
    @staticmethod
    def generate_transaction_history_excel(site_name, transactions_data, start_date=None, end_date=None):
        """
        Generate Excel report for transaction history
        """
        buffer = BytesIO()
        
        # Create DataFrame
        data = []
        for txn in transactions_data:
            data.append({
                'Serial Number': txn.serial_number,
                'Date': txn.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Type': txn.type.title(),
                'Material': txn.material_name,
                'Quantity': txn.quantity,
                'Unit': txn.unit,
                'Unit Cost': txn.unit_cost if txn.unit_cost else 0,
                'Total Value': txn.total_value if txn.total_value else 0,
                'Project Code': txn.project_code or 'N/A',
                'Created By': txn.created_by or 'System',
                'Notes': txn.notes or 'N/A'
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main transaction data sheet
            df.to_excel(writer, sheet_name='Transaction History', index=False)
            
            # Summary sheet
            if not df.empty:
                summary_data = {
                    'Metric': [
                        'Total Transactions', 
                        'Total Receipts', 
                        'Total Issues', 
                        'Total Adjustments',
                        'Total Value Impact',
                        'Date Range'
                    ],
                    'Value': [
                        len(df),
                        len(df[df['Type'] == 'Receipt']),
                        len(df[df['Type'] == 'Issue']),
                        len(df[df['Type'] == 'Adjustment']),
                        df['Total Value'].sum(),
                        f"{start_date or 'All'} to {end_date or 'All'}"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Material breakdown sheet
                material_summary = df.groupby('Material').agg({
                    'Quantity': 'sum',
                    'Total Value': 'sum'
                }).reset_index()
                material_summary.to_excel(writer, sheet_name='Material Summary', index=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format transaction history sheet
            worksheet = writer.sheets['Transaction History']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format currency columns
            from openpyxl.styles import NamedStyle
            currency_style = NamedStyle(name="currency")
            currency_style.number_format = '"ZMW" #,##0.00'
            
            # Apply currency formatting to cost and value columns
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                if row[6].value is not None:  # Unit Cost column
                    row[6].style = currency_style
                if row[7].value is not None:  # Total Value column
                    row[7].style = currency_style
        
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data
    
    @staticmethod
    def generate_daily_issues_excel(site_name, report_date, issues_data):
        """
        Generate Excel report for daily material issues
        """
        buffer = BytesIO()
        
        # Create DataFrame
        data = []
        logging.info(f"Processing {len(issues_data)} daily issues for Excel report")
        
        for issue in issues_data:
            data.append({
                'Serial Number': issue.serial_number,
                'Time': issue.created_at.strftime('%H:%M:%S'),
                'Material': issue.material_name,
                'Quantity': issue.quantity,
                'Unit': issue.unit,
                'Unit Cost': issue.unit_cost if issue.unit_cost else 0,
                'Total Value': issue.total_value if issue.total_value else 0,
                'Project Code': issue.issued_to_project_code or 'N/A',
                'Notes': issue.notes or 'N/A'
            })
        
        df = pd.DataFrame(data)
        logging.info(f"Created daily issues DataFrame with {len(df)} rows")
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main issues data sheet
            df.to_excel(writer, sheet_name='Daily Issues', index=False)
            
            # Summary sheet
            if not df.empty:
                summary_data = {
                    'Metric': [
                        'Report Date',
                        'Site Name', 
                        'Total Issues',
                        'Total Quantity',
                        'Total Value',
                        'Unique Materials'
                    ],
                    'Value': [
                        report_date.strftime('%Y-%m-%d'),
                        site_name,
                        len(df),
                        df['Quantity'].sum(),
                        df['Total Value'].sum(),
                        df['Material'].nunique()
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format daily issues sheet
            worksheet = writer.sheets['Daily Issues']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format currency columns
            from openpyxl.styles import NamedStyle
            currency_style = NamedStyle(name="currency")
            currency_style.number_format = '"ZMW" #,##0.00'
            
            # Apply currency formatting to cost and value columns
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                if row[5].value is not None:  # Unit Cost column
                    row[5].style = currency_style
                if row[6].value is not None:  # Total Value column
                    row[6].style = currency_style
        
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data